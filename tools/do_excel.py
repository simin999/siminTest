from openpyxl import load_workbook
from tools.get_path import *
from tools.read_config import ReadConfig
from tools.get_data import GetData
from tools.do_re import DoRegx
#从配置文件中读取要执行的sheet和用例
cf=eval(ReadConfig().read_config(config_path,'MODE','mode'))
class DoExcle:
    def __init__(self,filename):
        self.filename=filename

    def get_headers(self,sheetname):
        "获取表头字段（请求需要用到的，不需要写回结果的列）"
        wb = load_workbook(self.filename)
        sheet=wb[sheetname]
        headers=[]
        for j in range(1,sheet.max_column-2):
            headers.append(sheet.cell(1,j).value)
        return headers
    def get_data(self):
        "获取存储在excel表格中的测试数据"
        wb = load_workbook(self.filename)
        data = []
        #每次循环获取一个sheet的数据
        for key in cf:
            sheetname=key
            sheet=wb[sheetname]
            #定义一张表的数据
            sheet_data=[]
            #如果配置文件读取到的某个sheet需要执行的用例为all，读取此文件的所有数据
            if cf[key]=='all':
                #一次循环读取一行数据
                for i in range(2,sheet.max_row+1):
                    #定义行数据
                    row_data={}
                    #一次循环读取一个单元格，并根据表头字段赋值
                    for j in range(1,sheet.max_column-2):
                        row_data[self.get_headers(sheetname)[j-1]]=sheet.cell(i,j).value
                    #行数据的 sheetname 取值为当前读取的表格的sheetname
                    row_data['sheetname']=key
                    # 将已经产生确定值的变量批量替换，需要执行用例才产生确定值的变量不能替换
                    row_data = DoRegx.do_regx('\$\{(.*?)\}', str(row_data))
                    #将获取的每一行数据追加在表数据列表中
                    sheet_data.append(eval(row_data))
            else:
                for case_id in cf[key]:
                    row_data={}
                    for j in range(1,sheet.max_column-2):
                        row_data[self.get_headers(sheetname)[j-1]] = sheet.cell(case_id+1, j).value
                    row_data['sheetname'] = key
                    # 将已经产生确定值的变量批量替换（提前准备的测试数据做变量替换），需要执行用例后才产生确定值的变量（由测试数据执行后才生成的）不能替换
                    row_data=DoRegx.do_regx('\$\{(.*?)\}',str(row_data))
                    sheet_data.append(eval(row_data))
            data+=sheet_data
        return data

    def write_back(self,sheetname,i,j,value):
        "测试结果写回excel"
        wb = load_workbook(self.filename)
        sheet=wb[sheetname]
        sheet.cell(i,j).value=value
        wb.save(self.filename)
        wb.close()
if __name__ == '__main__':
    headers=data=DoExcle(test_data_path).get_headers("usedCar_collection")
    data=DoExcle(test_data_path).get_data()
    print(type(data))
    print(cf)
